import os
from flask import Flask, render_template, request, send_file
import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill

current_dir = os.path.abspath(os.path.dirname(__file__))
template_dir = os.path.join(current_dir, 'templates')
app = Flask(__name__, template_folder=template_dir)

def check_brand_presence(keyword, brand, marketplace="in"):
    url = f"https://www.amazon.{marketplace}/s?k={keyword.replace(' ', '+')}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')

        all_results = soup.find_all('div', {'data-component-type': 's-search-result'})

        sponsored_present = False
        organic_present = False

        for result in all_results:
            title_element = result.find('span', {'class': 'a-text-normal'})
            if not title_element:
                continue

            title = title_element.text
            is_sponsored = result.find('span', string=re.compile('Sponsored', re.IGNORECASE)) is not None

            if brand.lower() in title.lower():
                if is_sponsored:
                    sponsored_present = True
                else:
                    organic_present = True

            if sponsored_present and organic_present:
                break

        return sponsored_present, organic_present

    except requests.RequestException as e:
        print(f"Error fetching results: {e}")
        return None, None

def create_excel_file(results, brand_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon Search Results"

    # Add headers
    ws['A1'] = "Keyword"
    ws['B1'] = "Ads"
    ws['C1'] = "Organic"

    # Define colors
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Add data and apply colors
    for row, result in enumerate(results, start=2):
        ws.cell(row=row, column=1, value=result['keyword'])
        
        if result['sponsored_present']:
            ws.cell(row=row, column=2, value="Present").fill = green_fill
        else:
            ws.cell(row=row, column=2, value="Not Present").fill = red_fill

        if result['organic_present']:
            ws.cell(row=row, column=3, value="Present").fill = green_fill
        else:
            ws.cell(row=row, column=3, value="Not Present").fill = red_fill

    filename = f"amazon_search_results_{brand_name}.xlsx"
    wb.save(filename)
    return filename

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        brand_name = request.form['brand_name']
        keywords = request.form['keywords'].split('\n')
        results = []

        for keyword in keywords:
            keyword = keyword.strip()
            if keyword:
                sponsored_present, organic_present = check_brand_presence(keyword, brand_name)
                results.append({
                    'keyword': keyword,
                    'sponsored_present': sponsored_present,
                    'organic_present': organic_present
                })

        excel_file = create_excel_file(results, brand_name)
        return send_file(excel_file, as_attachment=True)

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
