import os
import logging
import random
import time
from flask import Flask, request, send_file, jsonify
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from flask_cors import CORS
import tempfile

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=True)
logging.basicConfig(level=logging.DEBUG)

def check_brand_presence(keyword, brand, marketplace="in"):
    url = f"https://www.amazon.in/s?k={keyword.replace(' ', '+')}&language=en_IN&currency=INR"
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15"
    ]
    headers = {
        "User-Agent": random.choice(user_agents),
        "Accept-Language": "en-IN,en-GB;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept-Currency": "INR"
    }
    
    session = requests.Session()
    retry = Retry(total=5, backoff_factor=0.1, status_forcelist=[500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    
    try:
        logging.info(f"Sending request to {url}")
        time.sleep(5)  # Wait for 5 seconds before each request
        response = session.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        logging.info("Response received successfully")
        
        soup = BeautifulSoup(response.content, 'html.parser')
        all_results = soup.find_all('div', {'data-component-type': 's-search-result'})
        logging.info(f"Found {len(all_results)} search results")
        
        sponsored_present = False
        organic_present = False
        sponsored_count = 0
        
        for result in all_results:
            title_element = result.find('span', {'class': 'a-text-normal'})
            if not title_element:
                continue
            title = title_element.text.strip()
            
            # Modified sponsored detection logic
            is_sponsored = (
                result.find('span', {'class': 'a-size-small a-color-base', 'dir': 'auto'}) and 
                result.find('span', text=re.compile('Sponsored'))
            ) is not None
            
            logging.info(f"Checking title: {title}, Sponsored: {is_sponsored}")
            
            if is_sponsored:
                sponsored_count += 1
                if brand.lower() in title.lower():
                    sponsored_present = True
                    logging.info("Sponsored presence detected")
                if sponsored_count >= 10:
                    break
            elif brand.lower() in title.lower():
                organic_present = True
                logging.info("Organic presence detected")
            
            if sponsored_present and organic_present:
                break
        
        logging.info(f"Final result - Sponsored: {sponsored_present}, Organic: {organic_present}")
        return sponsored_present, organic_present
    except requests.RequestException as e:
        logging.error(f"Error fetching results: {e}", exc_info=True)
        return None, None

def create_excel_file(results, brand_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon Search Results"
    
    # Add headers
    ws['A1'] = "Keyword"
    ws['B1'] = "Ads"
    ws['C1'] = "Organic"
    
    # Define colors
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Add data and apply colors
    for row, result in enumerate(results, start=2):
        ws.cell(row=row, column=1, value=result['keyword'])
        
        if result['sponsored_present']:
            ws.cell(row=row, column=2, value="Present").fill = green_fill
        else:
            ws.cell(row=row, column=2, value="Not Present").fill = red_fill
        
        if result['organic_present']:
            ws.cell(row=row, column=3, value="Present").fill = green_fill
        else:
            ws.cell(row=row, column=3, value="Not Present").fill = red_fill
    
    # Save to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        return tmp.name

@app.route('/api/check-brand', methods=['POST'])
def check_brand():
    try:
        logging.info("Received request to /api/check-brand")
        data = request.json
        logging.info(f"Received data: {data}")
        
        if not data:
            logging.error("No JSON data received")
            raise ValueError("No JSON data received")
        
        brand_name = data.get('brand_name')
        keywords = data.get('keywords')
        
        logging.info(f"Brand name: {brand_name}")
        logging.info(f"Keywords: {keywords}")
        
        if not brand_name or not keywords:
            logging.error("Missing brand_name or keywords")
            raise ValueError("Missing brand_name or keywords")

        results = []
        for keyword in keywords:
            keyword = keyword.strip()
            if keyword:
                logging.info(f"Checking keyword: {keyword}")
                sponsored_present, organic_present = check_brand_presence(keyword, brand_name)
                logging.info(f"Results for {keyword}: Sponsored: {sponsored_present}, Organic: {organic_present}")
                results.append({
                    'keyword': keyword,
                    'sponsored_present': sponsored_present,
                    'organic_present': organic_present
                })
        
        logging.info("Creating Excel file")
        excel_file = create_excel_file(results, brand_name)
        logging.info("Excel file created successfully")
        
        return send_file(excel_file, as_attachment=True, download_name=f"amazon_search_results_{brand_name}.xlsx")
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    try:
        app.run(host='0.0.0.0', port=port)
    except Exception as e:
        logging.error(f"An error occurred while starting the server: {str(e)}", exc_info=True)
